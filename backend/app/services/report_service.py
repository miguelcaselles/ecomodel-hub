"""
Report Generation Service
Handles PDF and Excel export of simulation results
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List
import io

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF generation with reportlab
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


class ReportService:
    """Service for generating reports in PDF and Excel formats"""

    def __init__(self):
        # Setup Jinja2 environment
        template_dir = Path(__file__).parent.parent / "templates"
        self.jinja_env = Environment(
            loader=FileSystemLoader(str(template_dir)),
            autoescape=select_autoescape(['html', 'xml'])
        )

        # Setup PDF styles
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom paragraph styles for PDF with HERA Value branding"""
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=26,
            spaceAfter=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#6366f1'),  # HERA Value indigo
            fontName='Helvetica-Bold'
        ))

        self.styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=self.styles['Heading2'],
            fontSize=15,
            spaceBefore=20,
            spaceAfter=12,
            textColor=colors.HexColor('#0f172a'),  # HERA Value slate
            fontName='Helvetica-Bold'
        ))

        self.styles.add(ParagraphStyle(
            name='Subtitle',
            parent=self.styles['Normal'],
            fontSize=13,
            textColor=colors.HexColor('#64748b'),  # HERA Value slate-500
            alignment=TA_CENTER,
            spaceAfter=20,
            fontName='Helvetica-Oblique'
        ))

        self.styles.add(ParagraphStyle(
            name='BrandName',
            parent=self.styles['Heading1'],
            fontSize=28,
            spaceAfter=8,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#6366f1'),  # HERA Value indigo
            fontName='Helvetica-Bold'
        ))

        self.styles.add(ParagraphStyle(
            name='CostEffective',
            parent=self.styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor('#059669'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))

        self.styles.add(ParagraphStyle(
            name='NotCostEffective',
            parent=self.styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor('#dc2626'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))

    def generate_pdf_report(
        self,
        scenario_name: str,
        user_email: str,
        organization: str,
        parameters: Dict[str, Any],
        results_drug_a: Dict[str, float],
        results_drug_b: Dict[str, float],
        psa_results: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Generate PDF report from simulation results using reportlab

        Args:
            scenario_name: Name of the scenario
            user_email: Email of user generating report
            organization: Organization name
            parameters: Model parameters
            results_drug_a: Results for Drug A
            results_drug_b: Results for Drug B
            psa_results: Optional PSA results

        Returns:
            bytes: PDF file content
        """
        # Calculate key metrics
        delta_costs = results_drug_a['total_cost'] - results_drug_b['total_cost']
        delta_qalys = results_drug_a['total_qalys'] - results_drug_b['total_qalys']

        if delta_qalys > 0:
            icer = delta_costs / delta_qalys
        else:
            icer = float('inf') if delta_costs > 0 else 0

        wtp_threshold = parameters.get('wtp_threshold', 30000)
        is_cost_effective = icer <= wtp_threshold

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Build story (content)
        story = []

        # Professional Header with HERA Value branding
        story.append(Paragraph("HERA Value®", self.styles['BrandName']))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("Cost-Effectiveness Analysis Report", self.styles['ReportTitle']))
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph(f"<i>{scenario_name}</i>", self.styles['Subtitle']))
        story.append(Spacer(1, 0.4*inch))

        # Add a horizontal line separator
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceAfter=0.3*inch))

        # Metadata table
        metadata = [
            ['Organization:', organization or 'N/A'],
            ['Generated by:', user_email],
            ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Time Horizon:', f"{parameters.get('time_horizon', 'N/A')} years"],
            ['Discount Rate:', f"{parameters.get('discount_rate', 0)}%"],
        ]

        meta_table = Table(metadata, colWidths=[3*cm, 8*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.4*inch))

        # Key Results Section
        story.append(Paragraph("Key Results", self.styles['SectionTitle']))

        results_data = [
            ['Metric', 'Drug A', 'Drug B', 'Difference'],
            ['Total Costs (€)', f"{results_drug_a['total_cost']:,.0f}",
             f"{results_drug_b['total_cost']:,.0f}", f"{delta_costs:,.0f}"],
            ['Total QALYs', f"{results_drug_a['total_qalys']:.3f}",
             f"{results_drug_b['total_qalys']:.3f}", f"{delta_qalys:.3f}"],
            ['Life Years', f"{results_drug_a.get('life_years', 0):.2f}",
             f"{results_drug_b.get('life_years', 0):.2f}",
             f"{results_drug_a.get('life_years', 0) - results_drug_b.get('life_years', 0):.2f}"],
        ]

        results_table = Table(results_data, colWidths=[4*cm, 3.5*cm, 3.5*cm, 3.5*cm])
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),  # HERA Value indigo
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(results_table)
        story.append(Spacer(1, 0.3*inch))

        # ICER Result
        story.append(Paragraph("Incremental Cost-Effectiveness Ratio (ICER)", self.styles['SectionTitle']))

        icer_display = f"€{icer:,.0f}/QALY" if icer != float('inf') else "Dominated"
        icer_data = [
            ['ICER', icer_display],
            ['WTP Threshold', f"€{wtp_threshold:,.0f}/QALY"],
        ]

        icer_table = Table(icer_data, colWidths=[5*cm, 5*cm])
        icer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ]))
        story.append(icer_table)
        story.append(Spacer(1, 0.2*inch))

        # Conclusion
        if is_cost_effective:
            conclusion = f"✓ Drug A is COST-EFFECTIVE at the €{wtp_threshold:,.0f}/QALY threshold"
            story.append(Paragraph(conclusion, self.styles['CostEffective']))
        else:
            conclusion = f"✗ Drug A is NOT COST-EFFECTIVE at the €{wtp_threshold:,.0f}/QALY threshold"
            story.append(Paragraph(conclusion, self.styles['NotCostEffective']))

        story.append(Spacer(1, 0.4*inch))

        # Parameters Section
        story.append(Paragraph("Model Parameters", self.styles['SectionTitle']))

        params_data = [
            ['Parameter', 'Value'],
            ['Drug A Cost (€/cycle)', f"€{parameters.get('cost_drug_a', 0):,.0f}"],
            ['Drug B Cost (€/cycle)', f"€{parameters.get('cost_drug_b', 0):,.0f}"],
            ['Healthcare Cost - Stable (€)', f"€{parameters.get('cost_stable', 0):,.0f}"],
            ['Healthcare Cost - Progression (€)', f"€{parameters.get('cost_progression', 0):,.0f}"],
            ['Progression Risk - Drug A', f"{parameters.get('prob_progression_a', 0)}%"],
            ['Progression Risk - Drug B', f"{parameters.get('prob_progression_b', 0)}%"],
            ['Utility - Stable', f"{parameters.get('utility_stable', 0):.2f}"],
            ['Utility - Progression', f"{parameters.get('utility_progression', 0):.2f}"],
        ]

        params_table = Table(params_data, colWidths=[8*cm, 6*cm])
        params_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b7280')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(params_table)

        # PSA Results (if available)
        if psa_results:
            story.append(Spacer(1, 0.4*inch))
            story.append(Paragraph("Probabilistic Sensitivity Analysis (PSA)", self.styles['SectionTitle']))

            percentiles = psa_results.get('percentiles', {})
            psa_data = [
                ['Metric', 'Value'],
                ['Number of Iterations', f"{psa_results.get('n_iterations', 1000):,}"],
                ['Mean ICER', f"€{psa_results.get('mean_icer', 0):,.0f}/QALY"],
                ['Median ICER (P50)', f"€{percentiles.get('p50', 0):,.0f}/QALY"],
                ['95% CI Lower (P2.5)', f"€{percentiles.get('p2_5', 0):,.0f}/QALY"],
                ['95% CI Upper (P97.5)', f"€{percentiles.get('p97_5', 0):,.0f}/QALY"],
                ['Probability Cost-Effective', f"{psa_results.get('prob_cost_effective', 0) * 100:.1f}%"],
            ]

            psa_table = Table(psa_data, colWidths=[7*cm, 7*cm])
            psa_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            story.append(psa_table)

        # Footer
        story.append(Spacer(1, 0.5*inch))
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceBefore=0.1*inch, spaceAfter=0.2*inch))
        footer_text = "<b>HERA Value®</b> - Professional Health Economic Analysis Platform"
        story.append(Paragraph(footer_text, self.styles['Subtitle']))

        # Build PDF
        doc.build(story)

        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel_report(
        self,
        scenario_name: str,
        user_email: str,
        parameters: Dict[str, Any],
        results_drug_a: Dict[str, float],
        results_drug_b: Dict[str, float],
        psa_results: Optional[Dict[str, Any]] = None,
        tornado_results: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Generate Excel report with multiple sheets

        Args:
            scenario_name: Name of the scenario
            user_email: Email of user generating report
            parameters: Model parameters
            results_drug_a: Results for Drug A
            results_drug_b: Results for Drug B
            psa_results: Optional PSA results
            tornado_results: Optional Tornado results

        Returns:
            bytes: Excel file content
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Define common styles with HERA Value branding
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")  # HERA Value indigo
        title_font = Font(bold=True, size=14, color="0F172A")  # HERA Value slate
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(
            ws_summary, scenario_name, user_email, parameters,
            results_drug_a, results_drug_b,
            header_font, header_fill, title_font, border
        )

        # Sheet 2: Parameters
        ws_params = wb.create_sheet("Parameters")
        self._create_parameters_sheet(
            ws_params, parameters,
            header_font, header_fill, title_font, border
        )

        # Sheet 3: Results
        ws_results = wb.create_sheet("Results")
        self._create_results_sheet(
            ws_results, results_drug_a, results_drug_b,
            header_font, header_fill, title_font, border
        )

        # Sheet 4: PSA (if available)
        if psa_results:
            ws_psa = wb.create_sheet("PSA")
            self._create_psa_sheet(
                ws_psa, psa_results,
                header_font, header_fill, title_font, border
            )

        # Sheet 5: Tornado (if available)
        if tornado_results:
            ws_tornado = wb.create_sheet("Tornado")
            self._create_tornado_sheet(
                ws_tornado, tornado_results,
                header_font, header_fill, title_font, border
            )

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()

    def _create_summary_sheet(self, ws, scenario_name, user_email, parameters,
                               results_drug_a, results_drug_b,
                               header_font, header_fill, title_font, border):
        """Create summary sheet with key metrics"""
        # Brand and Title
        ws['A1'] = 'HERA Value®'
        ws['A1'].font = Font(bold=True, size=18, color='6366F1')
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = 'Cost-Effectiveness Analysis Report'
        ws['A2'].font = Font(bold=True, size=14)
        ws.merge_cells('A2:D2')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Metadata
        row = 4
        ws[f'A{row}'] = 'Scenario:'
        ws[f'B{row}'] = scenario_name
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Generated By:'
        ws[f'B{row}'] = user_email
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Date:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)

        # Key Results
        row += 2
        ws[f'A{row}'] = 'KEY RESULTS'
        ws[f'A{row}'].font = title_font
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        # Header
        for col, header in enumerate(['Metric', 'Drug A', 'Drug B', 'Difference'], start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data rows
        delta_costs = results_drug_a['total_cost'] - results_drug_b['total_cost']
        delta_qalys = results_drug_a['total_qalys'] - results_drug_b['total_qalys']
        icer = delta_costs / delta_qalys if delta_qalys > 0 else 0

        data_rows = [
            ('Total Costs (€)', results_drug_a['total_cost'], results_drug_b['total_cost'], delta_costs),
            ('Total QALYs', results_drug_a['total_qalys'], results_drug_b['total_qalys'], delta_qalys),
            ('Life Years', results_drug_a.get('life_years', 0), results_drug_b.get('life_years', 0),
             results_drug_a.get('life_years', 0) - results_drug_b.get('life_years', 0)),
            ('ICER (€/QALY)', '', '', icer),
        ]

        for data_row in data_rows:
            row += 1
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                if isinstance(value, (int, float)) and value != '':
                    cell.number_format = '#,##0.00' if 'QALY' in data_row[0] or 'Years' in data_row[0] else '#,##0'

        # Conclusion
        row += 2
        wtp_threshold = parameters.get('wtp_threshold', 30000)
        is_cost_effective = icer <= wtp_threshold
        conclusion = f"Cost-Effective at €{wtp_threshold:,.0f}/QALY" if is_cost_effective else f"NOT Cost-Effective at €{wtp_threshold:,.0f}/QALY"

        ws[f'A{row}'] = 'Conclusion:'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = conclusion
        ws[f'B{row}'].font = Font(bold=True, size=12, color="059669" if is_cost_effective else "DC2626")
        ws.merge_cells(f'B{row}:D{row}')

        # Auto-fit columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_parameters_sheet(self, ws, parameters, header_font, header_fill, title_font, border):
        """Create parameters sheet"""
        ws['A1'] = 'Model Parameters'
        ws['A1'].font = title_font
        ws.merge_cells('A1:C1')

        row = 3
        # Header
        for col, header in enumerate(['Category', 'Parameter', 'Value'], start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Costs
        params_data = [
            ('Costs', 'Drug A Cost (€/cycle)', parameters.get('cost_drug_a', 0)),
            ('Costs', 'Drug B Cost (€/cycle)', parameters.get('cost_drug_b', 0)),
            ('Costs', 'Healthcare Cost - Stable (€)', parameters.get('cost_ae_drug_a', 0)),
            ('Costs', 'Healthcare Cost - Progression (€)', parameters.get('cost_progression', 0)),
            ('Clinical', 'Progression Risk - Drug A (%)', parameters.get('prob_progression_a', 0) * 100),
            ('Clinical', 'Progression Risk - Drug B (%)', parameters.get('prob_progression_b', 0) * 100),
            ('Clinical', 'Utility - Stable', parameters.get('utility_stable', 0)),
            ('Clinical', 'Utility - Progression', parameters.get('utility_progression', 0)),
            ('Settings', 'Time Horizon (years)', parameters.get('time_horizon', 0)),
            ('Settings', 'Discount Rate (%)', parameters.get('discount_rate', 0) * 100),
            ('Settings', 'WTP Threshold (€/QALY)', parameters.get('wtp_threshold', 0)),
        ]

        for data_row in params_data:
            row += 1
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                if col == 3 and isinstance(value, (int, float)):
                    if 'Utility' in data_row[1]:
                        cell.number_format = '0.00'
                    elif '%' in data_row[1]:
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '#,##0'

        # Auto-fit columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20

    def _create_results_sheet(self, ws, results_drug_a, results_drug_b, header_font, header_fill, title_font, border):
        """Create detailed results sheet"""
        ws['A1'] = 'Detailed Results'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        row = 3
        # Header
        for col, header in enumerate(['Metric', 'Drug A', 'Drug B', 'Difference'], start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # All results
        results_data = [
            ('Total Costs (€)', results_drug_a.get('total_cost', 0), results_drug_b.get('total_cost', 0)),
            ('Total QALYs', results_drug_a.get('total_qalys', 0), results_drug_b.get('total_qalys', 0)),
            ('Life Years', results_drug_a.get('life_years', 0), results_drug_b.get('life_years', 0)),
            ('Discounted Costs (€)', results_drug_a.get('discounted_costs', 0), results_drug_b.get('discounted_costs', 0)),
            ('Discounted QALYs', results_drug_a.get('discounted_qalys', 0), results_drug_b.get('discounted_qalys', 0)),
        ]

        for data_row in results_data:
            row += 1
            metric, val_a, val_b = data_row
            diff = val_a - val_b

            ws.cell(row=row, column=1).value = metric
            ws.cell(row=row, column=2).value = val_a
            ws.cell(row=row, column=3).value = val_b
            ws.cell(row=row, column=4).value = diff

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
                if col > 1:
                    if 'QALY' in metric or 'Years' in metric:
                        ws.cell(row=row, column=col).number_format = '0.000'
                    else:
                        ws.cell(row=row, column=col).number_format = '#,##0'

        # Auto-fit columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_psa_sheet(self, ws, psa_results, header_font, header_fill, title_font, border):
        """Create PSA results sheet"""
        ws['A1'] = 'Probabilistic Sensitivity Analysis (PSA)'
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')

        row = 3
        ws[f'A{row}'] = f"Number of Iterations: {psa_results.get('n_iterations', 1000)}"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:E{row}')

        row += 2
        # Header
        for col, header in enumerate(['Metric', 'Mean', 'P2.5', 'P50', 'P97.5'], start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # PSA data
        percentiles = psa_results.get('percentiles', {})
        row += 1
        ws.cell(row=row, column=1).value = 'ICER (€/QALY)'
        ws.cell(row=row, column=2).value = psa_results.get('mean_icer', 0)
        ws.cell(row=row, column=3).value = percentiles.get('p2_5', 0)
        ws.cell(row=row, column=4).value = percentiles.get('p50', 0)
        ws.cell(row=row, column=5).value = percentiles.get('p97_5', 0)

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            if col > 1:
                ws.cell(row=row, column=col).number_format = '#,##0'

        row += 2
        ws[f'A{row}'] = 'Probability Cost-Effective:'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{psa_results.get('prob_cost_effective', 0) * 100:.1f}%"
        ws[f'B{row}'].font = Font(bold=True, size=12)

        # Auto-fit columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_tornado_sheet(self, ws, tornado_results, header_font, header_fill, title_font, border):
        """Create Tornado analysis sheet"""
        ws['A1'] = 'Tornado Diagram - One-Way Sensitivity Analysis'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        row = 3
        # Header
        for col, header in enumerate(['Parameter', 'ICER Low', 'ICER High', 'Range'], start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Tornado data
        results = tornado_results.get('results', [])
        for item in results:
            row += 1
            ws.cell(row=row, column=1).value = item.get('parameter', '')
            ws.cell(row=row, column=2).value = item.get('low', 0)
            ws.cell(row=row, column=3).value = item.get('high', 0)
            ws.cell(row=row, column=4).value = abs(item.get('high', 0) - item.get('low', 0))

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
                if col > 1:
                    ws.cell(row=row, column=col).number_format = '#,##0'

        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        for col in range(2, 5):
            ws.column_dimensions[get_column_letter(col)].width = 15


    def generate_budget_impact_pdf(
        self,
        scenario_name: str,
        user_email: str,
        organization: str,
        parameters: Dict[str, Any],
        results: Dict[str, Any]
    ) -> bytes:
        """
        Generate PDF report for Budget Impact Analysis

        Args:
            scenario_name: Name of the scenario
            user_email: Email of user generating report
            organization: Organization name
            parameters: BIA parameters
            results: BIA results

        Returns:
            bytes: PDF file content
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Professional Header with HERA Value branding
        story.append(Paragraph("HERA Value®", self.styles['BrandName']))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("Budget Impact Analysis Report", self.styles['ReportTitle']))
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph(f"<i>{scenario_name}</i>", self.styles['Subtitle']))
        story.append(Spacer(1, 0.3*inch))

        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceAfter=0.3*inch))

        # Metadata
        metadata = [
            ['Organization:', organization or 'N/A'],
            ['Generated by:', user_email],
            ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Time Horizon:', f"{parameters.get('time_horizon', 'N/A')} years"],
            ['Target Population:', f"{parameters.get('target_population', 0):,} patients"],
        ]

        meta_table = Table(metadata, colWidths=[3*cm, 8*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.4*inch))

        # Executive Summary
        story.append(Paragraph("Executive Summary", self.styles['SectionTitle']))

        summary = results.get('summary', {})
        summary_data = [
            ['Metric', 'Value'],
            ['Total Budget Impact', f"€{summary.get('total_budget_impact', 0):,.0f}"],
            ['Budget Impact per Patient', f"€{summary.get('per_patient_impact', 0):,.0f}"],
            ['Years Until Budget Neutral', f"{summary.get('years_to_neutral', 'N/A')} years"],
        ]

        summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.4*inch))

        # Year-by-Year Budget Impact
        story.append(Paragraph("Year-by-Year Budget Impact", self.styles['SectionTitle']))

        yearly_data = [['Year', 'Current Scenario (€)', 'New Scenario (€)', 'Budget Impact (€)']]
        for year_result in results.get('yearly_results', []):
            yearly_data.append([
                f"Year {year_result.get('year', '')}",
                f"{year_result.get('current_budget', 0):,.0f}",
                f"{year_result.get('new_budget', 0):,.0f}",
                f"{year_result.get('budget_impact', 0):,.0f}"
            ])

        yearly_table = Table(yearly_data, colWidths=[3*cm, 4*cm, 4*cm, 4*cm])
        yearly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),  # HERA Value indigo
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(yearly_table)
        story.append(Spacer(1, 0.4*inch))

        # Market Share Assumptions
        story.append(Paragraph("Market Share Assumptions", self.styles['SectionTitle']))

        market_share = parameters.get('market_share_trajectory', [])
        if market_share:
            ms_data = [['Year', 'New Drug Market Share (%)']]
            for idx, share in enumerate(market_share, 1):
                ms_data.append([f"Year {idx}", f"{share}%"])

            ms_table = Table(ms_data, colWidths=[7*cm, 7*cm])
            ms_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b7280')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(ms_table)

        # Footer
        story.append(Spacer(1, 0.5*inch))
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceBefore=0.1*inch, spaceAfter=0.2*inch))
        footer_text = "<b>HERA Value®</b> - Professional Health Economic Analysis Platform"
        story.append(Paragraph(footer_text, self.styles['Subtitle']))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_decision_tree_pdf(
        self,
        scenario_name: str,
        user_email: str,
        organization: str,
        parameters: Dict[str, Any],
        results: Dict[str, Any]
    ) -> bytes:
        """
        Generate PDF report for Decision Tree Analysis
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Professional Header with HERA Value branding
        story.append(Paragraph("HERA Value®", self.styles['BrandName']))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("Decision Tree Analysis Report", self.styles['ReportTitle']))
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph(f"<i>{scenario_name}</i>", self.styles['Subtitle']))
        story.append(Spacer(1, 0.3*inch))

        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceAfter=0.3*inch))

        # Metadata
        metadata = [
            ['Organization:', organization or 'N/A'],
            ['Generated by:', user_email],
            ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Number of Strategies:', str(len(results.get('strategies', [])))],
        ]

        meta_table = Table(metadata, colWidths=[3*cm, 8*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.4*inch))

        # Strategy Results
        story.append(Paragraph("Strategy Comparison", self.styles['SectionTitle']))

        strategy_data = [['Strategy', 'Expected Cost (€)', 'Expected QALYs', 'Status']]
        for strategy in results.get('strategies', []):
            strategy_data.append([
                strategy.get('name', ''),
                f"{strategy.get('expected_cost', 0):,.0f}",
                f"{strategy.get('expected_qalys', 0):.3f}",
                strategy.get('status', '')
            ])

        strategy_table = Table(strategy_data, colWidths=[4*cm, 3.5*cm, 3.5*cm, 3.5*cm])
        strategy_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),  # HERA Value indigo
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(strategy_table)
        story.append(Spacer(1, 0.3*inch))

        # Optimal Strategy
        optimal = results.get('optimal_strategy', {})
        if optimal:
            story.append(Paragraph("Recommended Strategy", self.styles['SectionTitle']))
            story.append(Paragraph(
                f"<b>{optimal.get('name', 'N/A')}</b> - Expected Cost: €{optimal.get('expected_cost', 0):,.0f}, Expected QALYs: {optimal.get('expected_qalys', 0):.3f}",
                self.styles['CostEffective']
            ))

        # Footer
        story.append(Spacer(1, 0.5*inch))
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceBefore=0.1*inch, spaceAfter=0.2*inch))
        story.append(Paragraph("<b>HERA Value®</b> - Professional Health Economic Analysis Platform", self.styles['Subtitle']))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_survival_analysis_pdf(
        self,
        scenario_name: str,
        user_email: str,
        organization: str,
        parameters: Dict[str, Any],
        results: Dict[str, Any]
    ) -> bytes:
        """
        Generate PDF report for Survival Analysis
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Professional Header with HERA Value branding
        story.append(Paragraph("HERA Value®", self.styles['BrandName']))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("Parametric Survival Analysis Report", self.styles['ReportTitle']))
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph(f"<i>{scenario_name}</i>", self.styles['Subtitle']))
        story.append(Spacer(1, 0.3*inch))

        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceAfter=0.3*inch))

        # Metadata
        metadata = [
            ['Organization:', organization or 'N/A'],
            ['Generated by:', user_email],
            ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Analysis Type:', 'Parametric Survival Modeling'],
        ]

        meta_table = Table(metadata, colWidths=[3*cm, 8*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.4*inch))

        # Model Fit Statistics
        story.append(Paragraph("Distribution Fit Statistics", self.styles['SectionTitle']))

        fits = results.get('distribution_fits', [])
        if fits:
            fit_data = [['Distribution', 'AIC', 'BIC', 'Log-Likelihood']]
            for fit in fits:
                fit_data.append([
                    fit.get('distribution', ''),
                    f"{fit.get('aic', 0):.2f}",
                    f"{fit.get('bic', 0):.2f}",
                    f"{fit.get('log_likelihood', 0):.2f}"
                ])

            fit_table = Table(fit_data, colWidths=[4*cm, 3.5*cm, 3.5*cm, 3.5*cm])
            fit_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),  # HERA Value indigo
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            story.append(fit_table)

        story.append(Spacer(1, 0.3*inch))

        # Best Fit
        best_fit = results.get('best_fit', {})
        if best_fit:
            story.append(Paragraph("Recommended Distribution", self.styles['SectionTitle']))
            story.append(Paragraph(
                f"<b>{best_fit.get('distribution', 'N/A')}</b> (AIC: {best_fit.get('aic', 0):.2f})",
                self.styles['CostEffective']
            ))

            story.append(Spacer(1, 0.3*inch))

            # Parameters
            params = best_fit.get('parameters', {})
            if params:
                story.append(Paragraph("Distribution Parameters", self.styles['SectionTitle']))
                param_data = [['Parameter', 'Value']]
                for param, value in params.items():
                    param_data.append([param, f"{value:.4f}"])

                param_table = Table(param_data, colWidths=[7*cm, 7*cm])
                param_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b7280')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                ]))
                story.append(param_table)

        # Footer
        story.append(Spacer(1, 0.5*inch))
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceBefore=0.1*inch, spaceAfter=0.2*inch))
        story.append(Paragraph("<b>HERA Value®</b> - Professional Health Economic Analysis Platform", self.styles['Subtitle']))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_markov_flexible_pdf(
        self,
        scenario_name: str,
        user_email: str,
        organization: str,
        parameters: Dict[str, Any],
        results: Dict[str, Any]
    ) -> bytes:
        """
        Generate PDF report for Flexible Markov Model Analysis
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Professional Header with HERA Value branding
        story.append(Paragraph("HERA Value®", self.styles['BrandName']))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("Flexible Markov Model Analysis Report", self.styles['ReportTitle']))
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph(f"<i>{scenario_name}</i>", self.styles['Subtitle']))
        story.append(Spacer(1, 0.3*inch))

        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceAfter=0.3*inch))

        # Metadata
        metadata = [
            ['Organization:', organization or 'N/A'],
            ['Generated by:', user_email],
            ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Number of States:', str(len(parameters.get('states', [])))],
            ['Time Horizon:', f"{parameters.get('time_horizon', 'N/A')} cycles"],
        ]

        meta_table = Table(metadata, colWidths=[3*cm, 8*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.4*inch))

        # Model Structure
        story.append(Paragraph("Model Structure", self.styles['SectionTitle']))

        states = parameters.get('states', [])
        if states:
            state_data = [['State Name', 'Type', 'Cost (€)', 'Utility']]
            for state in states:
                state_data.append([
                    state.get('name', ''),
                    state.get('type', ''),
                    f"{state.get('cost', 0):,.0f}",
                    f"{state.get('utility', 0):.2f}"
                ])

            state_table = Table(state_data, colWidths=[4*cm, 3*cm, 3.5*cm, 3.5*cm])
            state_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),  # HERA Value indigo
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            story.append(state_table)

        story.append(Spacer(1, 0.4*inch))

        # Results Comparison
        story.append(Paragraph("Strategy Comparison", self.styles['SectionTitle']))

        strategies = results.get('strategies', [])
        if strategies:
            results_data = [['Strategy', 'Total Cost (€)', 'Total QALYs', 'ICER (€/QALY)']]
            for strategy in strategies:
                results_data.append([
                    strategy.get('name', ''),
                    f"{strategy.get('total_cost', 0):,.0f}",
                    f"{strategy.get('total_qalys', 0):.3f}",
                    f"{strategy.get('icer', 0):,.0f}" if strategy.get('icer') else 'Ref'
                ])

            results_table = Table(results_data, colWidths=[4*cm, 3.5*cm, 3.5*cm, 3.5*cm])
            results_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            story.append(results_table)

        # Footer
        story.append(Spacer(1, 0.5*inch))
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceBefore=0.1*inch, spaceAfter=0.2*inch))
        story.append(Paragraph("<b>HERA Value®</b> - Professional Health Economic Analysis Platform", self.styles['Subtitle']))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_voi_analysis_pdf(
        self,
        scenario_name: str,
        user_email: str,
        organization: str,
        parameters: Dict[str, Any],
        results: Dict[str, Any]
    ) -> bytes:
        """
        Generate PDF report for Value of Information Analysis
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Professional Header with HERA Value branding
        story.append(Paragraph("HERA Value®", self.styles['BrandName']))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("Value of Information Analysis Report", self.styles['ReportTitle']))
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph(f"<i>{scenario_name}</i>", self.styles['Subtitle']))
        story.append(Spacer(1, 0.3*inch))

        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceAfter=0.3*inch))

        # Metadata
        metadata = [
            ['Organization:', organization or 'N/A'],
            ['Generated by:', user_email],
            ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Analysis Type:', 'EVPI & EVPPI'],
        ]

        meta_table = Table(metadata, colWidths=[3*cm, 8*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.4*inch))

        # EVPI Results
        story.append(Paragraph("Expected Value of Perfect Information (EVPI)", self.styles['SectionTitle']))

        evpi = results.get('evpi', {})
        evpi_data = [
            ['Metric', 'Value'],
            ['EVPI per Patient', f"€{evpi.get('per_patient', 0):,.0f}"],
            ['Population EVPI', f"€{evpi.get('population', 0):,.0f}"],
            ['Decision Uncertainty', f"{evpi.get('decision_uncertainty', 0)*100:.1f}%"],
        ]

        evpi_table = Table(evpi_data, colWidths=[8*cm, 6*cm])
        evpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(evpi_table)
        story.append(Spacer(1, 0.4*inch))

        # EVPPI Results
        story.append(Paragraph("Expected Value of Partial Perfect Information (EVPPI)", self.styles['SectionTitle']))

        evppi_list = results.get('evppi', [])
        if evppi_list:
            evppi_data = [['Parameter', 'EVPPI per Patient (€)', 'Population EVPPI (€)', 'Priority']]
            for item in evppi_list:
                evppi_data.append([
                    item.get('parameter', ''),
                    f"{item.get('per_patient', 0):,.0f}",
                    f"{item.get('population', 0):,.0f}",
                    item.get('priority', '')
                ])

            evppi_table = Table(evppi_data, colWidths=[4*cm, 4*cm, 4*cm, 3*cm])
            evppi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),  # HERA Value indigo
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            story.append(evppi_table)

        # Interpretation
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("Interpretation", self.styles['SectionTitle']))
        interpretation = Paragraph(
            "The EVPI represents the maximum value of conducting further research to eliminate all uncertainty. "
            "The EVPPI shows which parameters would be most valuable to research further.",
            self.styles['Normal']
        )
        story.append(interpretation)

        # Footer
        story.append(Spacer(1, 0.5*inch))
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb'), spaceBefore=0.1*inch, spaceAfter=0.2*inch))
        story.append(Paragraph("<b>HERA Value®</b> - Professional Health Economic Analysis Platform", self.styles['Subtitle']))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()


# Global instance
report_service = ReportService()
